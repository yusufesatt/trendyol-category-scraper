from json import loads
from bs4 import BeautifulSoup as bs
from requests import get
import jpype
import asposecells
import pandas as pd
import openpyxl
import os

try:
        url = input("Trendyol Kategori Url'si giriniz: ")

        response = bs(get(url).text, "html.parser").find_all("script")
        result = response[12].text

        result = result.split("""window.__SEARCH_APP_INITIAL_STATE__=""")[-1]
        result = result.split(""";window.slpName='';window.TYPageName='product_search_result';window.isSearchResult=true;window.pageType="search";""")[0]
        result = loads(result)


        urun_isimleri = [name["name"] for name in result["products"]]
        urun_linkleri = [url["url"] for url in result["products"]]
        urun_resimleri = [img["images"] for img in result["products"]]
        urun_fiyat = [str(fiyat["variants"][0]["price"]["sellingPrice"])+" TL" for fiyat in result["products"]]

        urun_listesi = []

        for i in range(len(urun_isimleri)):
            urun_listesi.append([urun_isimleri[i], "https://cdn.dsmcdn.com"+urun_resimleri[i][0],"https://trendyol.com"+urun_linkleri[i], urun_fiyat[i]])

        df1 = pd.DataFrame(urun_listesi, columns=(["Ürün İsimleri", "Ürün Resimleri", "Ürün Linkleri", "Ürün Fiyatı"]))
        df1.to_excel("urunler.xlsx", sheet_name="urun", index=False)

        # Excel'de tabloları sığdırır
        jpype.startJVM()
        from asposecells.api import Workbook
        wb = Workbook("urunler.xlsx")
        worksheet = wb.getWorksheets().get(0)
        for i in range(4):
            worksheet.autoFitColumn(i)
        wb.save("urunler.xlsx")

        # Excelde reklam sayfasını siler
        wb = openpyxl.load_workbook('urunler.xlsx')
        wb.sheetnames
        std=wb['Evaluation Warning']
        wb.remove(std)
        wb.save('urunler.xlsx')

        print("Başarıyla excel'e aktarıldı :)")
except:
        print("Url'yi kontrol ediniz!")
